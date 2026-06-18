"""Advanced stock screener with multi-filter support, CSV and XLSX exports."""

import csv
import io
from datetime import datetime

from fastapi import APIRouter, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import func as sqlfunc, select

from app.core.deps import DbSession
from app.models.stock import Stock
from app.schemas.screener import ScreenerResult
from app.schemas.stock import StockRow

router = APIRouter()

_SORT_COLS = {
    "symbol": Stock.symbol,
    "last_price": Stock.last_price,
    "change_percent": Stock.change_percent,
    "volume": Stock.volume,
    "market_cap": Stock.market_cap,
    "pe_ratio": Stock.pe_ratio,
    "dividend_yield": Stock.dividend_yield,
}


def _build_query(
    sector: str | None,
    pe_min: float | None,
    pe_max: float | None,
    cap_min: float | None,
    cap_max: float | None,
    vol_min: float | None,
    div_yield_min: float | None,
    change_pct_min: float | None,
    change_pct_max: float | None,
    week52_pct_from_high: float | None,
):
    q = select(Stock)
    if sector:
        q = q.where(Stock.sector == sector)
    if pe_min is not None:
        q = q.where(Stock.pe_ratio != None, Stock.pe_ratio >= pe_min)  # noqa: E711
    if pe_max is not None:
        q = q.where(Stock.pe_ratio != None, Stock.pe_ratio <= pe_max)  # noqa: E711
    if cap_min is not None:
        q = q.where(Stock.market_cap != None, Stock.market_cap >= cap_min)  # noqa: E711
    if cap_max is not None:
        q = q.where(Stock.market_cap != None, Stock.market_cap <= cap_max)  # noqa: E711
    if vol_min is not None:
        q = q.where(Stock.volume != None, Stock.volume >= vol_min)  # noqa: E711
    if div_yield_min is not None:
        q = q.where(Stock.dividend_yield != None, Stock.dividend_yield >= div_yield_min)  # noqa: E711
    if change_pct_min is not None:
        q = q.where(Stock.change_percent != None, Stock.change_percent >= change_pct_min)  # noqa: E711
    if change_pct_max is not None:
        q = q.where(Stock.change_percent != None, Stock.change_percent <= change_pct_max)  # noqa: E711
    if week52_pct_from_high is not None:
        # Price within `week52_pct_from_high`% below its 52-week high
        # last_price >= week52_high * (1 - week52_pct_from_high/100)
        q = q.where(
            Stock.week52_high != None,  # noqa: E711
            Stock.last_price != None,  # noqa: E711
            Stock.last_price >= Stock.week52_high * (1 - week52_pct_from_high / 100),
        )
    return q


@router.get("", response_model=ScreenerResult)
async def screen(
    db: DbSession,
    sector: str | None = None,
    pe_min: float | None = None,
    pe_max: float | None = None,
    cap_min: float | None = None,
    cap_max: float | None = None,
    vol_min: float | None = None,
    div_yield_min: float | None = None,
    change_pct_min: float | None = None,
    change_pct_max: float | None = None,
    week52_pct_from_high: float | None = None,
    sort_by: str = Query("market_cap", pattern="^(symbol|last_price|change_percent|volume|market_cap|pe_ratio|dividend_yield)$"),
    sort_dir: str = Query("desc", pattern="^(asc|desc)$"),
    limit: int = Query(100, le=500),
    offset: int = Query(0, ge=0),
) -> ScreenerResult:
    base_q = _build_query(sector, pe_min, pe_max, cap_min, cap_max, vol_min, div_yield_min, change_pct_min, change_pct_max, week52_pct_from_high)

    total = await db.scalar(select(sqlfunc.count()).select_from(base_q.subquery())) or 0

    col = _SORT_COLS.get(sort_by, Stock.market_cap)
    order = col.asc() if sort_dir == "asc" else col.desc().nulls_last()
    stocks = list((await db.scalars(base_q.order_by(order).limit(limit).offset(offset))).all())

    return ScreenerResult(total=total, stocks=[StockRow.model_validate(s) for s in stocks])


@router.get("/export/csv")
async def export_csv(
    db: DbSession,
    sector: str | None = None,
    pe_min: float | None = None,
    pe_max: float | None = None,
    cap_min: float | None = None,
    cap_max: float | None = None,
    vol_min: float | None = None,
    div_yield_min: float | None = None,
    change_pct_min: float | None = None,
    change_pct_max: float | None = None,
    week52_pct_from_high: float | None = None,
    sort_by: str = Query("market_cap"),
    sort_dir: str = Query("desc"),
) -> StreamingResponse:
    base_q = _build_query(sector, pe_min, pe_max, cap_min, cap_max, vol_min, div_yield_min, change_pct_min, change_pct_max, week52_pct_from_high)
    col = _SORT_COLS.get(sort_by, Stock.market_cap)
    order = col.asc() if sort_dir == "asc" else col.desc().nulls_last()
    stocks = list((await db.scalars(base_q.order_by(order).limit(2000))).all())

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Symbol", "Name", "Sector", "Price (NGN)", "Change %", "Volume", "Market Cap", "P/E", "Div Yield %", "52W High", "52W Low"])
    for s in stocks:
        writer.writerow([
            s.symbol, s.name, s.sector or "",
            s.last_price or "", s.change_percent or "",
            s.volume or "", s.market_cap or "",
            s.pe_ratio or "", s.dividend_yield or "",
            s.week52_high or "", s.week52_low or "",
        ])

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=ngx_screener.csv"},
    )


@router.get("/export/xlsx")
async def export_xlsx(
    db: DbSession,
    sector: str | None = None,
    pe_min: float | None = None,
    pe_max: float | None = None,
    cap_min: float | None = None,
    cap_max: float | None = None,
    vol_min: float | None = None,
    div_yield_min: float | None = None,
    change_pct_min: float | None = None,
    change_pct_max: float | None = None,
    week52_pct_from_high: float | None = None,
    sort_by: str = Query("market_cap"),
    sort_dir: str = Query("desc"),
) -> StreamingResponse:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        from fastapi import HTTPException, status
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail="openpyxl not installed — XLSX export unavailable",
        )

    base_q = _build_query(sector, pe_min, pe_max, cap_min, cap_max, vol_min,
                          div_yield_min, change_pct_min, change_pct_max, week52_pct_from_high)
    col = _SORT_COLS.get(sort_by, Stock.market_cap)
    order = col.asc() if sort_dir == "asc" else col.desc().nulls_last()
    stocks = list((await db.scalars(base_q.order_by(order).limit(5000))).all())

    wb = Workbook()
    ws = wb.active
    ws.title = "NGX Screener"

    # ── Palette ──
    TEAL = "1A6B5A"
    LIGHT_TEAL = "E6F3EF"
    MID_GRAY = "F5F5F5"
    DARK_TEXT = "1A1A1A"
    GAIN_COLOR = "15803D"
    LOSS_COLOR = "B91C1C"
    NEUTRAL = "6B7280"

    header_fill = PatternFill("solid", fgColor=TEAL)
    alt_fill = PatternFill("solid", fgColor=LIGHT_TEAL)
    gray_fill = PatternFill("solid", fgColor=MID_GRAY)
    thin_border = Border(
        bottom=Side(style="thin", color="D1D5DB"),
    )

    # ── Title row ──
    ws.merge_cells("A1:K1")
    title_cell = ws["A1"]
    title_cell.value = "Vortex · NGX Screener Export"
    title_cell.font = Font(bold=True, size=14, color="FFFFFF")
    title_cell.fill = PatternFill("solid", fgColor=TEAL)
    title_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 28

    # ── Subtitle / filter summary ──
    ws.merge_cells("A2:K2")
    filters_used = []
    if sector:             filters_used.append(f"Sector: {sector}")
    if pe_min is not None: filters_used.append(f"P/E ≥ {pe_min}")
    if pe_max is not None: filters_used.append(f"P/E ≤ {pe_max}")
    if cap_min is not None: filters_used.append(f"Cap ≥ ₦{cap_min/1e9:.1f}B")
    if cap_max is not None: filters_used.append(f"Cap ≤ ₦{cap_max/1e9:.1f}B")
    if vol_min is not None: filters_used.append(f"Vol ≥ {vol_min:,.0f}")
    if div_yield_min is not None: filters_used.append(f"Div yield ≥ {div_yield_min}%")
    if change_pct_min is not None: filters_used.append(f"Chg ≥ {change_pct_min}%")
    if change_pct_max is not None: filters_used.append(f"Chg ≤ {change_pct_max}%")
    sub_cell = ws["A2"]
    filter_text = "  ·  ".join(filters_used) if filters_used else "No filters applied"
    sub_cell.value = f"{len(stocks)} results  |  {filter_text}  |  Exported {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    sub_cell.font = Font(size=10, color="FFFFFF", italic=True)
    sub_cell.fill = PatternFill("solid", fgColor="2D8B70")
    sub_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 18

    ws.row_dimensions[3].height = 6  # spacer

    # ── Headers (row 4) ──
    COLS = [
        ("Symbol",      10, "symbol"),
        ("Name",        28, "name"),
        ("Sector",      16, "sector"),
        ("Price (₦)",   12, "last_price"),
        ("Change %",    10, "change_percent"),
        ("Volume",      13, "volume"),
        ("Mkt Cap (₦)", 16, "market_cap"),
        ("P/E",          8, "pe_ratio"),
        ("Div Yield %",  11, "dividend_yield"),
        ("52W High",    12, "week52_high"),
        ("52W Low",     12, "week52_low"),
    ]
    for col_idx, (label, width, _) in enumerate(COLS, start=1):
        cell = ws.cell(row=4, column=col_idx, value=label)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[4].height = 20

    # ── Data rows ──
    for row_idx, s in enumerate(stocks, start=5):
        is_alt = (row_idx % 2 == 0)
        row_fill = alt_fill if is_alt else None
        chg = s.change_percent or 0.0

        values = [
            s.symbol,
            s.name,
            s.sector or "",
            s.last_price,
            s.change_percent,
            s.volume,
            s.market_cap,
            s.pe_ratio,
            s.dividend_yield,
            s.week52_high,
            s.week52_low,
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            if row_fill:
                cell.fill = row_fill

            # Number formatting
            if col_idx == 4:  # price
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right")
            elif col_idx == 5:  # change %
                cell.number_format = '+0.00%;-0.00%'
                cell.alignment = Alignment(horizontal="right")
                if val is not None:
                    cell.value = (val or 0) / 100
                    cell.font = Font(
                        color=GAIN_COLOR if chg >= 0 else LOSS_COLOR,
                        bold=True,
                    )
            elif col_idx in (6, 7):  # volume, market cap
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal="right")
            elif col_idx in (8, 9, 10, 11):  # pe, div, 52w
                cell.number_format = '0.00'
                cell.alignment = Alignment(horizontal="right")
            elif col_idx == 1:  # symbol
                cell.font = Font(bold=True)
        ws.row_dimensions[row_idx].height = 16

    # ── Freeze header rows ──
    ws.freeze_panes = "A5"

    # ── Save ──
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"ngx_screener_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
